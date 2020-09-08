# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 18:10:10 2020

@author: ll

Modified on Tue Sep 8 10:13:22 2020
@Stephen Qian
"""

import  pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

def str2num(x):
    if x=='':
        return 0
    else:
        return int(x)

def SumAndFormat(paraout, df=None):
    workbook = load_workbook(filename=paraout)
    sheet = workbook['Details'] #.active
    sheet.delete_rows(idx=3)
    sheet.delete_cols(idx=1)
    if df is None:
        data=sheet.values
        columns = next(data)[0:]
        df = pd.DataFrame(data,columns=columns)
    # df = df.reset_index()

    maxVal=[['']*3,['']*3]

    sizeO = df.loc[:,('Size of Parameters','SizeO')].apply(str2num)
    sizeW = df.loc[:,('Size of Parameters','SizeW')].apply(str2num)
    opGEMM = df.loc[:,('Operation Summary','GEMM')].apply(str2num)

    # max size
    maxVal[0][0] = sizeO.max() # activation
    maxVal[0][1] = sizeW.max()
    # max ops
    maxVal[1][0] = opGEMM.max()

    sumlst = [
        ['Total Activations(MB):',sizeO.sum()/(1000**2)],
        ['Total Weights(MB):',sizeW.sum()/(1000**2)],
        ['Total GEMM (G_ops):',opGEMM.sum()/(1000**3)]
    ]

    SumSheetGen(paraout, sumlst)
    # FormatTable(paraout, maxVal)

def SumSheetGen(paraout, sumlst):
    workbook = load_workbook(filename=paraout)
    if "Summary" in set(workbook.sheetnames):
        sheet = workbook["Summary"]
    else:
        sheet = workbook.create_sheet("Summary")
    for i in range(1,4):
        sheet["A{}".format(i)] = sumlst[i-1][0]
        sheet["B{}".format(i)] = sumlst[i-1][1]

    sheet.insert_rows(idx=1)
    summarystr = 'Model Statistics:'
    sheet['A1']= summarystr
    sheet['A1'].font=Font(b=True)
    sheet.column_dimensions['A'].width = max(25,len(summarystr))
    workbook.save(paraout)


def FormatTable(paraout, maxVal):
    maxSiActi = maxVal[0][0]
    maxSiWeig = maxVal[0][1]
    maxOpGemm = maxVal[1][0]

    workbook = load_workbook(filename=paraout)
    sheet = workbook['Details']#.active
    sheet.freeze_panes = "C2"

    # row 0: Grey bkcolor, Bold font
    for cell in list(sheet.rows)[0]:
        cell.fill = PatternFill("solid", fgColor="00C0C0C0")
        cell.font = Font(b=True)
        if cell.value=='SizeO':
            so=cell.column_letter
        if cell.value=='SizeW':
            sw=cell.column_letter
        if cell.value=='OpGemm':
            og=cell.column_letter

    # Max activation row with red
    background = PatternFill(bgColor="00FF0000")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxSiActi)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(so+'{}:'.format(sheet.min_row)+so+'{}'.format(sheet.max_row), myrule)

    # Max activation row with pink
    background = PatternFill(bgColor="00FF00FF")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxSiWeig)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(sw+'{}:'.format(sheet.min_row)+sw+'{}'.format(sheet.max_row), myrule)

    #  Max Ops Gemm row with green
    background = PatternFill(bgColor="0000FF00")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxOpGemm)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(og+'{}:'.format(sheet.min_row)+og+'{}'.format(sheet.max_row), myrule)

    workbook.save(paraout)